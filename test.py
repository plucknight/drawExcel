from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os

def image_to_excel(
    image_path,
    output_file=None,
    max_columns=100,  # 限制最大列数（防止Excel卡死）
    cell_width_ratio=0.3  # Excel默认单元格宽高比
):
    # 自动生成输出文件名
    if output_file is None:
        base_name = os.path.splitext(os.path.basename(image_path))[0]
        output_file = f"res/{base_name}.xlsx"

    # 读取图片并计算缩放尺寸
    img = Image.open(image_path).convert("RGB")
    orig_width, orig_height = img.size
    orig_ratio = orig_width / orig_height

    # 根据Excel单元格宽高比反向补偿缩放比例
    target_ratio = orig_ratio * cell_width_ratio  # 目标宽高比 = 原图比例 * 单元格宽高比

    # 计算缩放后的尺寸（不超过最大列数）
    if orig_width > max_columns:
        new_width = max_columns
        new_height = int(new_width / target_ratio)
    else:
        new_width = orig_width
        new_height = int(orig_height / cell_width_ratio)  # 反向补偿高度

    # 保持比例缩放
    img = img.resize((new_width, new_height))

    # 创建Excel工作簿（不调整列宽和行高）
    wb = Workbook()
    ws = wb.active

    # 遍历像素填充颜色
    for y in range(new_height):
        for x in range(new_width):
            r, g, b = img.getpixel((x, y))
            color_hex = f"{r:02x}{g:02x}{b:02x}"
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
            cell = ws.cell(row=y + 1, column=x + 1)
            cell.fill = fill

    # 保存文件
    wb.save(output_file)
    print(f"生成成功！尺寸：{new_width}x{new_height}，文件：{output_file}")

if __name__ == "__main__":
    image_path = "img/sh.jpg"
    image_to_excel(image_path)
